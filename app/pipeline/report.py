import os
from typing import Dict, Any, List
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook


def _env(template_dir: str) -> Environment:
    return Environment(
        loader=FileSystemLoader(template_dir),
        autoescape=select_autoescape(["html", "xml"]),
    )


def generate_html_report(context: Dict[str, Any], template_dir: str, template_name: str, out_path: str) -> str:
    env = _env(template_dir)
    html = env.get_template(template_name).render(**context)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def _write_sheet(ws, rows: List[List[Any]]):
    for row in rows:
        ws.append(row)


def generate_excel_report(context: Dict[str, Any], out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()

    # Emissions sheet
    ws = wb.active
    ws.title = "Emissions"
    emissions = context.get("emissions", {})
    _write_sheet(ws, [["Metric", "Value"], *[[k, v] for k, v in emissions.items()]])

    # Fields sheet
    ws2 = wb.create_sheet("Fields")
    fields = context.get("fields", {})
    _write_sheet(ws2, [["Field", "Value"], *[[k, v] for k, v in fields.items()]])

    # Flags sheet
    ws3 = wb.create_sheet("Flags")
    flags = context.get("flags", [])
    _write_sheet(ws3, [["Severity", "Message", "Field", "Value"]])
    for f in flags:
        ws3.append([f.get("severity"), f.get("message"), f.get("field"), f.get("value", "")])

    wb.save(out_path)
    return out_path